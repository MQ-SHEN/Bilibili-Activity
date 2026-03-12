import os
import time
import json
import threading
import subprocess
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService

# --- 环境与路径初始化 ---
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def install_and_import(package, import_name=None):
    if import_name is None: import_name = package
    try:
        return __import__(import_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        return __import__(import_name)

# 自动安装必要依赖
webdriver_manager = install_and_import("webdriver-manager", "webdriver_manager")
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
install_and_import("tkcalendar")
from tkcalendar import Calendar
install_and_import("openpyxl")


# ===================== 自定义样式增强版日期选择对话框 =====================
class DatePickerDialog(tk.Toplevel):
    def __init__(self, parent, initial_date=None):
        super().__init__(parent)
        self.title("选择日期")
        self.result = None
        
        # 强制置顶并锁定父窗口
        self.transient(parent)
        self.grab_set()
        
        # 1. 调整窗口大小以确保按钮完整显示 (宽 350, 高 320)
        w, h = 350, 320
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (w // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.resizable(False, False) # 固定大小防止变形

        # 2. 配置日历样式
        # 通过 headersfont 设置年月标题的大小
        # 通过 selectbackground 设置选中颜色
        self.cal = Calendar(self, selectmode='day', 
                           year=initial_date.year if initial_date else datetime.now().year,
                           month=initial_date.month if initial_date else datetime.now().month,
                           day=initial_date.day if initial_date else datetime.now().day,
                           locale='zh_CN', 
                           date_pattern='yyyy-mm-dd',
                           # 关键样式定制：
                           showweeknumbers=False,      # 隐藏周数，腾出空间
                           headersfont=('微软雅黑', 10, 'bold'), # 年份月份字体变大
                           font=('微软雅黑', 10),        # 日期数字字体
                           background='#0078D4',       # 头部背景色
                           foreground='white',         # 头部文字颜色
                           selectbackground='#FF9900', # 选中日期颜色
                           # 年份月份位置说明：Calendar默认年份左上，月份右上
                           # 如果需要更极端的左右分离，可以通过下面的 headersbackgroud 调整
                           )
        self.cal.pack(padx=10, pady=(15, 10), fill=tk.BOTH, expand=True)

        # 3. 按钮区域布局优化
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 使用统一风格的按钮
        style = ttk.Style()
        style.configure("Active.TButton", font=("微软雅黑", 10, "bold"))

        self.ok_btn = ttk.Button(btn_frame, text="确 定", command=self.on_ok, style="Active.TButton")
        self.ok_btn.pack(side=tk.LEFT, padx=(50, 0), ipadx=10, ipady=5)
        
        self.cancel_btn = ttk.Button(btn_frame, text="取 消", command=self.destroy)
        self.cancel_btn.pack(side=tk.RIGHT, padx=(0, 50), ipadx=10, ipady=5)

        # 绑定回车键直接确认
        self.bind('<Return>', lambda e: self.on_ok())

    def on_ok(self):
        self.result = self.cal.get_date()
        self.destroy()

def parse_date_str(s):
    if not s: return None
    s = s.strip()
    fmts = ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%Y年%m月%d日"]
    for f in fmts:
        try: return datetime.strptime(s, f).date()
        except: pass
    return None

# ===================== 爬虫逻辑类 =====================
class BilibiliCrawler:
    def __init__(self, status_callback=None):
        self.status_callback = status_callback
        self.driver = None
        self.is_running = False

    def update_status(self, msg):
        if self.status_callback: self.status_callback(msg)

    def get_driver_service(self, browser, manual_path, auto_match):
        if not auto_match:
            if not manual_path or not os.path.exists(manual_path):
                raise ValueError("请先手动选择驱动路径，或开启自动匹配！")
            if browser == "Chrome": return ChromeService(manual_path)
            elif browser == "Firefox": return FirefoxService(manual_path)
            return EdgeService(manual_path)

        try:
            self.update_status(f"正在匹配 {browser} 驱动版本...")
            if browser == "Chrome": path = ChromeDriverManager().install()
            elif browser == "Firefox": path = GeckoDriverManager().install()
            else: path = EdgeChromiumDriverManager().install()
            self.update_status(f"✅ 驱动匹配成功")
            if browser == "Chrome": return ChromeService(path)
            elif browser == "Firefox": return FirefoxService(path)
            return EdgeService(path)
        except Exception as e:
            self.update_status(f"❌ 自动匹配失败: {str(e)}")
            raise

    def crawl(self, browser, manual_path, auto_match, save_path, pages, prog_var, start_date_str):
        self.is_running = True
        start_date = parse_date_str(start_date_str)
        try:
            service = self.get_driver_service(browser, manual_path, auto_match)
            if browser == "Chrome":
                opts = webdriver.ChromeOptions()
                opts.add_experimental_option('excludeSwitches', ['enable-logging'])
                self.driver = webdriver.Chrome(service=service, options=opts)
            elif browser == "Firefox":
                self.driver = webdriver.Firefox(service=service)
            else:
                self.driver = webdriver.Edge(service=service)

            results = []
            for p in range(1, pages + 1):
                if not self.is_running: break
                self.update_status(f"正在读取第 {p}/{pages} 页...")
                self.driver.get(f"https://www.bilibili.com/blackboard/activity-list.html?page={p}")
                time.sleep(2)
                
                items = self.driver.find_elements(By.CSS_SELECTOR, ".act-list li")
                for li in items:
                    try:
                        status = li.find_element(By.CSS_SELECTOR, ".event_status span").text.strip()
                        if "已结束" in status: continue
                        
                        full_text = li.find_element(By.CSS_SELECTOR, ".event_status").text.strip()
                        date_str = full_text.replace(status, "").strip()
                        s_date = date_str.split("至")[0].strip() if "至" in date_str else date_str
                        
                        p_start = parse_date_str(s_date)
                        if start_date and p_start and p_start < start_date: continue
                        
                        a_tag = li.find_element(By.CSS_SELECTOR, "h2 a")
                        results.append([a_tag.text.strip(), a_tag.get_attribute("href"), status, s_date, 
                                       date_str.split("至")[1].strip() if "至" in date_str else ""])
                    except: continue
                prog_var.set(int((p / pages) * 100))

            if results:
                self.save_excel(results, save_path)
                return True
            return False
        finally:
            self.is_running = False
            if self.driver: self.driver.quit()

    def save_excel(self, data, path):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook(); ws = wb.active; ws.title = "B站活动列表"
        headers = ["活动标题(点击打开)", "链接", "状态", "开始日期", "结束日期"]
        ws.append(headers)
        h_fill = PatternFill("solid", fgColor="0078D4"); h_font = Font(color="FFFFFF", bold=True)
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i); c.fill = h_fill; c.font = h_font; c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[c.column_letter].width = 45 if i <= 2 else 15
        for row in data:
            ws.append(row); curr = ws.max_row
            ws.cell(curr, 1).hyperlink = row[1]; ws.cell(curr, 1).font = Font(color="0000FF", underline="single")
        wb.save(path)
        self.update_status(f"✅ 文件已保存至目标目录")

# ===================== GUI 界面 =====================
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("B站活动自动化采集工具")
        
        self.window_width = 700
        self.window_height = 550
        self.center_window()
        self.set_app_icon()
        
        self.root.configure(bg="#f5f7f9")
        self.crawler = BilibiliCrawler(self.update_status)
        self.last_file = ""
        
        self.setup_ui()
        self.load_config()

    def resource_path(self, relative_path):
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base_path, relative_path)

    def center_window(self):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width // 2) - (self.window_width // 2)
        y = (screen_height // 2) - (self.window_height // 2)
        self.root.geometry(f"{self.window_width}x{self.window_height}+{x}+{y}")

    def set_app_icon(self):
        icon_path = self.resource_path("logo.ico")
        if os.path.exists(icon_path):
            try: self.root.iconbitmap(icon_path)
            except: pass

    def setup_ui(self):
        style = ttk.Style(); style.theme_use('clam')
        style.configure("TFrame", background="#f5f7f9")
        style.configure("TLabelframe", background="#ffffff")
        style.configure("TLabelframe.Label", background="#ffffff", font=("微软雅黑", 10, "bold"))
        
        container = ttk.Frame(self.root, padding=15)
        container.pack(fill=tk.BOTH, expand=True)

        cfg_frame = ttk.LabelFrame(container, text=" 任务设置 ", padding=12)
        cfg_frame.pack(fill=tk.X, pady=(0, 10))

        row1 = ttk.Frame(cfg_frame); row1.pack(fill=tk.X, pady=5)
        ttk.Label(row1, text="浏览器:", background="#fff").pack(side=tk.LEFT)
        self.browser_var = tk.StringVar(value="Chrome")
        ttk.Combobox(row1, textvariable=self.browser_var, values=["Chrome", "Edge", "Firefox"], width=8, state="readonly").pack(side=tk.LEFT, padx=5)
        
        ttk.Label(row1, text="页数:", background="#fff").pack(side=tk.LEFT, padx=(15, 0))
        self.page_var = tk.StringVar(value="1")
        ttk.Spinbox(row1, from_=1, to=100, textvariable=self.page_var, width=5).pack(side=tk.LEFT, padx=5)

        ttk.Label(row1, text="日期筛选 ≥", background="#fff").pack(side=tk.LEFT, padx=(15, 0))
        
        # --- 替换逻辑：Entry + 按钮弹出独立日历 ---
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.date_entry = ttk.Entry(row1, textvariable=self.date_var, width=12, state="readonly")
        self.date_entry.pack(side=tk.LEFT, padx=2)
        ttk.Button(row1, text="📅", width=3, command=self.pick_date).pack(side=tk.LEFT, padx=(0, 5))

        self.auto_match_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(row1, text="自动匹配驱动", variable=self.auto_match_var).pack(side=tk.LEFT, padx=(10, 0))

        row2 = ttk.Frame(cfg_frame); row2.pack(fill=tk.X, pady=8)
        ttk.Label(row2, text="驱动路径:", background="#fff").pack(side=tk.LEFT)
        self.driver_path_var = tk.StringVar()
        ttk.Entry(row2, textvariable=self.driver_path_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(row2, text="选择文件", command=self.pick_driver, width=8).pack(side=tk.LEFT)

        row3 = ttk.Frame(cfg_frame); row3.pack(fill=tk.X)
        ttk.Label(row3, text="保存位置:", background="#fff").pack(side=tk.LEFT)
        self.save_path_var = tk.StringVar()
        ttk.Entry(row3, textvariable=self.save_path_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(row3, text="选择目录", command=self.pick_save_dir, width=8).pack(side=tk.LEFT)

        ctrl_frame = ttk.Frame(container); ctrl_frame.pack(fill=tk.X, pady=10)
        self.run_btn = ttk.Button(ctrl_frame, text="▶ 开始执行", command=self.start_task, width=15)
        self.run_btn.pack(side=tk.LEFT, padx=5)
        self.open_btn = ttk.Button(ctrl_frame, text="📂 打开生成文件", command=self.open_last_file, state=tk.DISABLED)
        self.open_btn.pack(side=tk.LEFT, padx=5)
        self.stop_btn = ttk.Button(ctrl_frame, text="停止", command=self.stop_task, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.RIGHT, padx=5)

        self.prog_var = tk.IntVar()
        self.prog_bar = ttk.Progressbar(container, variable=self.prog_var, maximum=100)
        self.prog_bar.pack(fill=tk.X, pady=(0, 10))

        self.log_area = scrolledtext.ScrolledText(container, height=7, font=("Consolas", 9), bg="#ffffff", state=tk.DISABLED)
        self.log_area.pack(fill=tk.BOTH, expand=True)

    def pick_date(self):
        # 弹出独立的对话框窗口
        try:
            current = datetime.strptime(self.date_var.get(), "%Y-%m-%d")
        except:
            current = datetime.now()
            
        dialog = DatePickerDialog(self.root, initial_date=current)
        self.root.wait_window(dialog)
        if dialog.result:
            self.date_var.set(dialog.result)

    def pick_driver(self):
        p = filedialog.askopenfilename(filetypes=[("驱动文件", "*.exe")])
        if p: self.driver_path_var.set(p)

    def pick_save_dir(self):
        d = filedialog.askdirectory()
        if d:
            name = f"B站活动_{datetime.now().strftime('%m%d_%H%M')}.xlsx"
            self.save_path_var.set(os.path.join(d, name))

    def update_status(self, msg):
        self.log_area.config(state=tk.NORMAL)
        self.log_area.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_area.see(tk.END); self.log_area.config(state=tk.DISABLED)

    def open_last_file(self):
        if self.last_file and os.path.exists(self.last_file):
            if sys.platform == 'win32': os.startfile(self.last_file)
            else: subprocess.call(['open' if sys.platform=='darwin' else 'xdg-open', self.last_file])

    def start_task(self):
        if not self.save_path_var.get():
            messagebox.showwarning("提示", "请选择保存位置！")
            return
        self.run_btn.config(state=tk.DISABLED); self.stop_btn.config(state=tk.NORMAL)
        self.open_btn.config(state=tk.DISABLED); self.prog_var.set(0)
        self.save_config()
        threading.Thread(target=self.worker_thread, daemon=True).start()

    def worker_thread(self):
        try:
            success = self.crawler.crawl(
                self.browser_var.get(), self.driver_path_var.get(), self.auto_match_var.get(),
                self.save_path_var.get(), int(self.page_var.get()), self.prog_var, self.date_var.get()
            )
            if success:
                self.last_file = self.save_path_var.get()
                self.open_btn.config(state=tk.NORMAL)
                messagebox.showinfo("完成", "爬取成功！")
        except Exception as e:
            messagebox.showerror("错误", str(e))
        finally:
            self.run_btn.config(state=tk.NORMAL); self.stop_btn.config(state=tk.DISABLED)

    def stop_task(self):
        self.crawler.is_running = False
        self.update_status("正在强制停止...")

    def save_config(self):
        path = os.path.join(BASE_DIR, "crawler_settings.json")
        conf = {
            "browser": self.browser_var.get(),
            "driver_path": self.driver_path_var.get(),
            "auto_match": self.auto_match_var.get(),
            "save_path": self.save_path_var.get(),
            "date": self.date_var.get()
        }
        with open(path, "w", encoding="utf-8") as f: json.dump(conf, f, indent=2)

    def load_config(self):
        path = os.path.join(BASE_DIR, "crawler_settings.json")
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    c = json.load(f)
                    self.browser_var.set(c.get("browser", "Chrome"))
                    self.driver_path_var.set(c.get("driver_path", ""))
                    self.auto_match_var.set(c.get("auto_match", True))
                    self.save_path_var.set(c.get("save_path", ""))
                    self.date_var.set(c.get("date", datetime.now().strftime("%Y-%m-%d")))
            except: pass

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()